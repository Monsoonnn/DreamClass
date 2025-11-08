from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Data for Chapter 1
data = [
    ["ID", "Question", "A", "B", "C", "D", "Correct"],
    [1, "Dao động là gì?", "Chuyển động lặp lại quanh vị trí cân bằng.", "Chuyển động thẳng đều.", "Chuyển động tròn.", "Chuyển động rơi tự do.", "A"],
    [2, "Công thức li độ trong dao động điều hòa?", "x = A sin(ωt + φ)", "x = A cos(ωt + φ)", "x = A tan(ωt + φ)", "x = A cot(ωt + φ)", "B"],
    [3, "Năng lượng toàn phần trong dao động điều hòa?", "W = (1/2) k A^2", "W = (1/2) m v^2", "W = m g h", "W = F s", "A"],
    [4, "Điều kiện cộng hưởng trong dao động?", "f = f0", "f > f0", "f < f0", "f = 0", "A"],
    [5, "Chu kỳ dao động con lắc đơn?", "T = 2π sqrt(l/g)", "T = 2π sqrt(g/l)", "T = sqrt(2π l/g)", "T = sqrt(2π g/l)", "A"]
]

wb = Workbook()
ws = wb.active

# Write header
for col, value in enumerate(data[0], start=1):
    ws.cell(row=1, column=col, value=value)

# Write data
for row_num, row_data in enumerate(data[1:], start=2):
    for col, value in enumerate(row_data[:-1], start=1):
        ws.cell(row=row_num, column=col, value=value)
    
    # Color correct option
    correct = row_data[-1]
    col_map = {'A': 3, 'B': 4, 'C': 5, 'D': 6}
    correct_col = col_map[correct]
    fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    ws.cell(row=row_num, column=correct_col).fill = fill

wb.save("Chapter1.xlsx")
print("File Chapter1.xlsx đã được tạo với màu xanh cho đáp án đúng.")

# ĐƯA CHATGPT/GEMINI ĐỂ CHẠY 